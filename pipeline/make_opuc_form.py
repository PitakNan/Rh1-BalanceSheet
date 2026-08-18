# -*- coding: utf-8 -*-
"""
make_opuc_form.py — สร้างแบบกรอก (Excel) ให้จังหวัด/รพ. ระบุ "เจ้าหนี้–ลูกหนี้ค่ารักษา
OP-UC นอก CUP (ในจังหวัด) รายคู่โรงพยาบาล" ณ สิ้นเดือน  · เจ้าของงานสั่ง 12 ส.ค. 69

หัวใจของแบบฟอร์มนี้ 3 อย่าง (อย่าตัดออก — เป็นเหตุผลที่มันได้ข้อมูลที่ใช้งานได้จริง):
 ① **ยอดควบคุมจากงบทดลองของ รพ. เอง** — แต่ละบล็อกมีช่องยอดรวมตามบัญชี 2101020199.202 /
    1102050101.203+1102050194.204 ที่ รพ. ส่ง HFO ไว้แล้ว พร้อมสูตรเทียบผลรวมที่กรอก
    → ถ้ากรอกไม่ครบ/เกิน ช่องสถานะจะขึ้นแดงทันที ไม่ต้องรอเขตตรวจ
 ② **เติมยอดที่ระบบแกะได้ให้ล่วงหน้า** (จากชื่อบัญชีย่อยที่ รพ. ตั้งเอง) → รพ. ที่ทำอยู่แล้ว
    แค่ตรวจทาน ไม่ต้องกรอกใหม่ทั้งหมด
 ③ **คอลัมน์อ้างอิง "คู่สัญญาบันทึกไว้เท่าไหร่"** → เห็นความต่างระหว่างสองฝั่งตั้งแต่ตอนกรอก
    ซึ่งเป็นงานกระทบยอดตัวจริง (งวด 256910 คู่ที่ครบสองทางตรงกันเป๊ะแค่ 8/72 คู่)

ต้องรัน analyze_opuc_pairs.py ก่อน (ใช้ opuc_pairs.csv เป็นแหล่งยอดที่แกะได้)
ผลลัพธ์: pipeline/out/form_opuc_<งวด>.xlsx  — 1 แท็บคำชี้แจง + 1 แท็บต่อจังหวัด
"""
import io, os, csv, sys, json, collections
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

OUT_DIR = r"D:\Github\Rh1-BalanceSheet\pipeline\out"
EXEC_J  = r"D:\Github\Rh1-BalanceSheet\docs\data\risk\exec.json"
PAIRS   = os.path.join(OUT_DIR, "opuc_pairs.csv")
ASOF    = "31 กรกฎาคม 2569"
PERIOD  = "256910"

NAVY   = "1F4E79"
BLUE   = "4472C4"
YELLOW = "FFF2CC"      # ช่องกรอก
GRAY   = "F2F2F2"      # ช่องอ้างอิง ห้ามแก้
GREEN  = "E2EFDA"
THIN   = Side(style="thin", color="BFBFBF")
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAHT   = '#,##0.00'

def main():
    ex = json.load(io.open(EXEC_J, encoding="utf-8"))
    HOSP = {h["hcode"]: h for h in ex["hosp"]}
    prov_h = collections.defaultdict(list)
    for h in ex["hosp"]:
        prov_h[h.get("prov", "?")].append(h)
    # เรียง รพศ./รพท. ขึ้นก่อน แล้วตามชื่อ — ให้แม่ข่ายอยู่ต้นแท็บ (เป็นคู่สัญญาหลักของทุกแห่ง)
    order = {"รพศ.": 0, "รพท.": 1, "รพช.": 2}
    for p in prov_h:
        prov_h[p].sort(key=lambda h: (order.get(h.get("type"), 9), h["name"]))

    # ยอดที่ระบบแกะได้: (ผู้จ่าย, ผู้รับ) -> {'pay':x,'ar':y}
    got = {}
    if os.path.exists(PAIRS):
        for r in csv.DictReader(io.open(PAIRS, encoding="utf-8-sig")):
            k = (r["ผู้จ่าย(hcode)"], r["ผู้รับ(hcode)"])
            f = lambda s: float(s) if s not in ("", None) else None
            got[k] = {"pay": f(r["ยอดที่ผู้จ่ายบันทึก"]), "ar": f(r["ยอดที่ผู้รับบันทึก"])}

    wb = Workbook()
    # ══ แท็บคำชี้แจง ═══════════════════════════════════════════════════════
    ws = wb.active; ws.title = "คำชี้แจง"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108
    def line(txt, bold=False, size=11, fill=None, indent=0):
        r = ws.max_row + 1
        c = ws.cell(row=r, column=2, value=txt)
        c.font = Font(bold=bold, size=size, color="FFFFFF" if fill == NAVY else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=indent)
        if fill: c.fill = PatternFill("solid", fgColor=fill)
        # 🪤 ตั้งความสูงตายตัว 30 ไม่พอสำหรับข้อความ 3 บรรทัด → บรรทัดท้ายถูกตัด
        #    openpyxl auto-fit ไม่ได้ จึงคำนวณจากความยาว (คอลัมน์ B กว้าง ~105 อักษร)
        n = max(1, -(-len(str(txt)) // 100))
        ws.row_dimensions[r].height = 15 * n if n > 1 else None
        return r
    ws["B1"] = f"แบบกรอกเจ้าหนี้–ลูกหนี้ค่ารักษา OP-UC นอก CUP (ในจังหวัด) รายคู่โรงพยาบาล  ณ {ASOF}"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B1"].alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells("B1:E1")      # 🪤 ไม่ merge แล้วหัวเรื่อง 14pt ถูกตัดกลางคำ (เจอตอนตรวจภาพ)
    ws.row_dimensions[1].height = 30
    line("")
    line("วัตถุประสงค์", True, 12, GREEN)
    line("เขตสุขภาพที่ 1 ต้องการทราบว่า เจ้าหนี้/ลูกหนี้ค่ารักษาผู้ป่วยนอก UC ที่ไปรับบริการนอก CUP "
         "ในจังหวัดเดียวกัน ของโรงพยาบาลแต่ละแห่ง เป็นหนี้/ลูกหนี้ของโรงพยาบาลใด เป็นจำนวนเท่าไหร่ "
         "เพื่อใช้จัดเวทีกระทบยอดและหักกลบระหว่างหน่วยบริการภายในจังหวัด")
    line("")
    line("บัญชีที่เกี่ยวข้อง (ตามผังบัญชีกระทรวงสาธารณสุข)", True, 12, GREEN)
    line("• เจ้าหนี้  2101020199.202  เจ้าหนี้ค่ารักษา OP-UC นอก CUP (ในจังหวัดสังกัด สธ.)", indent=1)
    line("• ลูกหนี้  1102050101.203 และ 1102050194.204  ลูกหนี้ค่ารักษา UC-OP นอก CUP (ในจังหวัด)", indent=1)
    line("หมายเหตุ: แบบฟอร์มนี้ไม่รวมบัญชี 'ต่างจังหวัด' (2101020199.203 / 1102050101.204 / 1102050194.205) "
         "และไม่รวมหน่วยบริการนอกสังกัด สป.สธ.", indent=1)
    line("")
    line("วิธีกรอก", True, 12, GREEN)
    line("1) เปิดแท็บจังหวัดของท่าน — ในแท็บมีบล็อกของโรงพยาบาลทุกแห่งในจังหวัด เรียงจากแม่ข่ายลงมา", indent=1)
    line("2) กรอกเฉพาะ 2 ช่องสีเหลือง: 'เจ้าหนี้ – เราค้างจ่ายเขา' และ 'ลูกหนี้ – เขาค้างจ่ายเรา' "
         "เป็นยอดคงค้าง ณ วันสิ้นเดือน หน่วยบาท (ทศนิยม 2 ตำแหน่ง)", indent=1)
    line("3) ช่องที่มีเลขอยู่แล้ว = ระบบดึงมาจากชื่อบัญชีย่อยที่โรงพยาบาลตั้งไว้ในงบทดลองที่ส่ง HFO "
         "โปรด 'ตรวจทานและแก้ให้ถูกต้อง' ไม่ต้องกรอกใหม่ทั้งหมด", indent=1)
    line("4) ช่องสีเทาแก้ไม่ได้ — เป็นข้อมูลอ้างอิง: 'คู่สัญญาบันทึกไว้เท่าไหร่' ใช้ดูว่าสองฝั่งตรงกันหรือไม่", indent=1)
    line("5) ดูแถวสุดท้ายของแต่ละบล็อก: ระบบเทียบผลรวมที่กรอกกับยอดรวมตามงบทดลองให้อัตโนมัติ "
         "ถ้าขึ้น ❌ แปลว่ายังกรอกไม่ครบหรือเกิน ให้ตรวจอีกครั้งก่อนส่ง", indent=1)
    line("6) หากมีคู่สัญญาที่ไม่อยู่ในรายชื่อ (เช่น รพ.ในจังหวัดที่ไม่ได้สังกัด สป.สธ.) ให้ใช้แถว "
         "'อื่น ๆ (โปรดระบุ)' ท้ายบล็อก", indent=1)
    line("")
    line("ยอดควบคุม", True, 12, GREEN)
    line("ยอดรวมตามงบทดลองที่แสดงในแต่ละบล็อก มาจากงบทดลองงวดเดือนกรกฎาคม 2569 ที่โรงพยาบาล "
         "ส่งเข้าระบบ HFO แล้ว (ยอดเดียวกับที่เขตใช้ในแดชบอร์ดวิกฤตการเงิน) จึงไม่ควรแก้ "
         "ถ้าเห็นว่ายอดควบคุมไม่ถูกต้อง แปลว่าต้องแก้ที่งบทดลอง ไม่ใช่แก้ในแบบฟอร์มนี้")
    line("")
    line("สถานะปัจจุบันของแต่ละจังหวัด (ก่อนกรอกแบบฟอร์มนี้)", True, 12, GREEN)
    hdr = ["จังหวัด", "รพ.ทั้งหมด", "แตกบัญชีย่อยแล้ว (เจ้าหนี้)", "แตกบัญชีย่อยแล้ว (ลูกหนี้)"]
    r0 = ws.max_row + 1
    for i, t in enumerate(hdr):
        c = ws.cell(row=r0, column=2 + i, value=t)
        c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(wrap_text=True, horizontal="center"); c.border = BOX
    nsub = collections.defaultdict(lambda: [0, 0])
    for (a, b), v in got.items():
        pa = HOSP.get(a, {}).get("prov"); pb = HOSP.get(b, {}).get("prov")
        if v.get("pay") is not None and pa: nsub[pa][0] = nsub[pa][0]
    seen_pay = collections.defaultdict(set); seen_ar = collections.defaultdict(set)
    for (a, b), v in got.items():
        if v.get("pay") is not None: seen_pay[HOSP.get(a, {}).get("prov")].add(a)
        if v.get("ar") is not None:  seen_ar[HOSP.get(b, {}).get("prov")].add(b)
    for i, p in enumerate(sorted(prov_h, key=lambda x: -len(prov_h[x]))):
        r = r0 + 1 + i
        for j, val in enumerate([p, len(prov_h[p]), len(seen_pay.get(p, ())), len(seen_ar.get(p, ()))]):
            c = ws.cell(row=r, column=2 + j, value=val)
            c.border = BOX
            if j: c.alignment = Alignment(horizontal="center")
    for col, w in (("C", 12), ("D", 24), ("E", 24)):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

    # ══ แท็บรายจังหวัด ═════════════════════════════════════════════════════
    for prov in sorted(prov_h, key=lambda x: -len(prov_h[x])):
        hs = prov_h[prov]
        w = wb.create_sheet(prov[:31])
        w.sheet_view.showGridLines = False
        widths = [(1, 5), (2, 10), (3, 30), (4, 18), (5, 18), (6, 18), (7, 18), (8, 26)]
        for i, ww in widths: w.column_dimensions[get_column_letter(i)].width = ww
        w["A1"] = f"จังหวัด{prov} — เจ้าหนี้/ลูกหนี้ค่ารักษา OP-UC นอก CUP (ในจังหวัด) ณ {ASOF}"
        w["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        w.merge_cells("A1:H1")
        w["A1"].fill = PatternFill("solid", fgColor=NAVY)
        w["A1"].alignment = Alignment(vertical="center", indent=1)
        w.row_dimensions[1].height = 26
        # ⚠️ ข้อความยาวต้องอยู่บรรทัดนี้ (merge A2:H2 กว้างพอ) ไม่ใช่ในเซลล์สถานะที่ merge แค่ F:H
        #    ซึ่งกว้างไม่พอ แล้วข้อความจะถูกตัดกลางคำ (เจอจริงตอนตรวจภาพ 12 ส.ค. 69)
        w["A2"] = ("กรอกเฉพาะช่องสีเหลือง · ช่องสีเทาเป็นข้อมูลอ้างอิงห้ามแก้ · "
                   "ดูแถว 'ผลต่าง' ท้ายแต่ละบล็อกให้ขึ้น ✅ ก่อนส่ง "
                   "(ผลต่างติดลบ = ยังกรอกไม่ครบ · เป็นบวก = กรอกเกินยอดในบัญชี · "
                   "ยอมรับความต่างได้ไม่เกิน 1 บาท)")
        w["A2"].font = Font(size=10, italic=True, color="555555")
        w.merge_cells("A2:H2")
        row = 4
        for h in hs:
            hc = h["hcode"]
            others = [o for o in hs if o["hcode"] != hc]
            gl_pay = h.get("tj", {}).get("payIn", 0) or 0
            gl_ar  = h.get("tj", {}).get("arIn", 0) or 0
            # ── หัวบล็อก ──
            w.cell(row=row, column=1, value=f"{h['name']}  ({h['type']} · รหัส {hc})").font = \
                Font(bold=True, size=12, color="FFFFFF")
            w.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            w.cell(row=row, column=1).fill = PatternFill("solid", fgColor=BLUE)
            w.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
            w.row_dimensions[row].height = 22
            row += 1
            w.cell(row=row, column=1,
                   value=(f"ยอดรวมตามงบทดลอง ณ {ASOF}:   เจ้าหนี้ (2101020199.202) = "
                          f"{gl_pay:,.2f} บาท      ลูกหนี้ (1102050101.203 + 1102050194.204) = "
                          f"{gl_ar:,.2f} บาท")).font = Font(size=10, bold=True, color="1F4E79")
            w.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1
            # ── หัวตาราง ──
            heads = ["ลำดับ", "รหัส รพ.", "ชื่อโรงพยาบาลคู่สัญญา",
                     "เจ้าหนี้\nเราค้างจ่ายเขา (บาท)", "ลูกหนี้\nเขาค้างจ่ายเรา (บาท)",
                     "อ้างอิง: เขาบันทึกว่า\nเราค้างจ่ายเขา", "อ้างอิง: เขาบันทึกว่า\nเขาค้างจ่ายเรา",
                     "หมายเหตุ"]
            for i, t in enumerate(heads, start=1):
                c = w.cell(row=row, column=i, value=t)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="7F7F7F")
                c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                c.border = BOX
            w.row_dimensions[row].height = 30
            hrow = row
            row += 1
            first = row
            for i, o in enumerate(others, start=1):
                oc = o["hcode"]
                mine_pay = (got.get((hc, oc), {}) or {}).get("pay")     # เราตั้งเจ้าหนี้ให้เขา
                mine_ar  = (got.get((oc, hc), {}) or {}).get("ar")      # เราตั้งลูกหนี้จากเขา
                ref_pay  = (got.get((hc, oc), {}) or {}).get("ar")      # เขาตั้งลูกหนี้จากเรา
                ref_ar   = (got.get((oc, hc), {}) or {}).get("pay")     # เขาตั้งเจ้าหนี้ให้เรา
                w.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
                w.cell(row=row, column=2, value=oc).alignment = Alignment(horizontal="center")
                w.cell(row=row, column=3, value=f"{o['name']} ({o['type']})")
                for col, val in ((4, mine_pay), (5, mine_ar)):
                    c = w.cell(row=row, column=col, value=val)
                    c.fill = PatternFill("solid", fgColor=YELLOW); c.number_format = BAHT
                for col, val in ((6, ref_pay), (7, ref_ar)):
                    c = w.cell(row=row, column=col, value=val)
                    c.fill = PatternFill("solid", fgColor=GRAY); c.number_format = BAHT
                    c.font = Font(size=10, italic=True, color="7F7F7F")
                w.cell(row=row, column=8).fill = PatternFill("solid", fgColor=YELLOW)
                for col in range(1, 9): w.cell(row=row, column=col).border = BOX
                row += 1
            # แถวสำรองสำหรับคู่สัญญาที่ไม่อยู่ในรายชื่อ
            for k in range(2):
                w.cell(row=row, column=3, value="อื่น ๆ (โปรดระบุชื่อหน่วยบริการ)").font = \
                    Font(size=10, italic=True, color="7F7F7F")
                for col in (3, 4, 5, 8):
                    w.cell(row=row, column=col).fill = PatternFill("solid", fgColor=YELLOW)
                for col in (4, 5): w.cell(row=row, column=col).number_format = BAHT
                for col in range(1, 9): w.cell(row=row, column=col).border = BOX
                row += 1
            last = row - 1
            # ── รวม + ตรวจยอด ──
            w.cell(row=row, column=3, value="รวมที่กรอก").font = Font(bold=True)
            for col in (4, 5):
                c = w.cell(row=row, column=col,
                           value=f"=SUM({get_column_letter(col)}{first}:{get_column_letter(col)}{last})")
                c.font = Font(bold=True); c.number_format = BAHT
                c.fill = PatternFill("solid", fgColor=GREEN)
            for col in range(1, 9): w.cell(row=row, column=col).border = BOX
            sum_row = row
            row += 1
            w.cell(row=row, column=3, value="ผลต่าง = ที่กรอก − งบทดลอง").font = Font(bold=True)
            w.cell(row=row, column=3).alignment = Alignment(horizontal="right")
            for col, gl in ((4, gl_pay), (5, gl_ar)):
                cl = get_column_letter(col)
                c = w.cell(row=row, column=col, value=f"={cl}{sum_row}-{gl:.2f}")
                c.font = Font(bold=True); c.number_format = BAHT
            c = w.cell(row=row, column=6,
                       value=f'=IF(AND(ABS(D{row})<1,ABS(E{row})<1),"✅ ยอดตรงกับงบทดลองแล้ว",'
                             f'"❌ ยังไม่ตรง — ติดลบ = กรอกไม่ครบ")')
            c.font = Font(bold=True); c.alignment = Alignment(horizontal="left")
            w.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
            # แดงเมื่อยังไม่ตรง (ทำที่ช่องผลต่างทั้งสองช่อง)
            for col in (4, 5):
                cl = get_column_letter(col)
                w.conditional_formatting.add(
                    f"{cl}{row}", FormulaRule(formula=[f"ABS({cl}{row})>=1"],
                                              fill=PatternFill("solid", fgColor="FFC7CE"),
                                              font=Font(bold=True, color="9C0006")))
                w.conditional_formatting.add(
                    f"{cl}{row}", FormulaRule(formula=[f"ABS({cl}{row})<1"],
                                              fill=PatternFill("solid", fgColor="C6EFCE"),
                                              font=Font(bold=True, color="006100")))
            for col in range(1, 9): w.cell(row=row, column=col).border = BOX
            row += 3            # เว้นบรรทัดก่อนบล็อกถัดไป
        w.freeze_panes = "A4"

    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, f"form_opuc_{PERIOD}.xlsx")
    wb.save(path)
    print(f"เขียนแล้ว: {path}")
    print(f"  แท็บ: คำชี้แจง + {len(prov_h)} จังหวัด · รพ. รวม {sum(len(v) for v in prov_h.values())} แห่ง")
    print(f"  ช่องที่เติมยอดให้ล่วงหน้าได้: {sum(1 for v in got.values() if v.get('pay') is not None)} เจ้าหนี้ / "
          f"{sum(1 for v in got.values() if v.get('ar') is not None)} ลูกหนี้")

if __name__ == "__main__":
    main()
